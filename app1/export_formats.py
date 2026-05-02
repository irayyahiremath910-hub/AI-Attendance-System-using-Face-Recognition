"""
Advanced export formats for attendance data
Supports CSV, Excel, PDF, JSON, and custom formats
"""

import csv
import json
from datetime import date, datetime
from io import BytesIO, StringIO
import logging

logger = logging.getLogger(__name__)


class ExportService:
    """Core export service for multiple formats"""

    @staticmethod
    def export_to_csv(data, filename="attendance_export.csv"):
        """Export data to CSV format"""
        try:
            output = StringIO()
            
            if not data:
                return None

            # Get headers from first record
            if isinstance(data[0], dict):
                fieldnames = list(data[0].keys())
            else:
                fieldnames = ['value']

            writer = csv.DictWriter(output, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(data)

            csv_content = output.getvalue()
            logger.info(f"CSV export created: {filename}")
            return csv_content

        except Exception as e:
            logger.error(f"CSV export error: {str(e)}")
            return None

    @staticmethod
    def export_to_json(data, filename="attendance_export.json", pretty=True):
        """Export data to JSON format"""
        try:
            json_data = json.dumps(
                data,
                indent=2 if pretty else None,
                default=str,
                sort_keys=True
            )
            logger.info(f"JSON export created: {filename}")
            return json_data

        except Exception as e:
            logger.error(f"JSON export error: {str(e)}")
            return None

    @staticmethod
    def export_to_xlsx(data, filename="attendance_export.xlsx"):
        """Export data to Excel format"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Attendance"

            if not data:
                return None

            # Headers
            if isinstance(data[0], dict):
                headers = list(data[0].keys())
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")

                # Data rows
                for row_idx, record in enumerate(data, 2):
                    for col_idx, header in enumerate(headers, 1):
                        value = record.get(header, '')
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.value = value
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                # Auto-adjust column widths
                for col in range(1, len(headers) + 1):
                    ws.column_dimensions[get_column_letter(col)].width = 15

            # Save to bytes
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            logger.info(f"XLSX export created: {filename}")
            return output.getvalue()

        except ImportError:
            logger.error("openpyxl not installed")
            return None
        except Exception as e:
            logger.error(f"XLSX export error: {str(e)}")
            return None

    @staticmethod
    def export_to_pdf(data, filename="attendance_export.pdf", title="Attendance Report"):
        """Export data to PDF format"""
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.units import inch
            from reportlab.lib import colors

            output = BytesIO()
            doc = SimpleDocTemplate(output, pagesize=A4)
            elements = []

            # Title
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.HexColor('#1f4788'),
                spaceAfter=30,
                alignment=1  # Center
            )
            elements.append(Paragraph(title, title_style))
            elements.append(Spacer(1, 0.3 * inch))

            # Data table
            if data:
                headers = list(data[0].keys()) if isinstance(data[0], dict) else ['Data']
                table_data = [headers]

                for record in data:
                    if isinstance(record, dict):
                        row = [str(record.get(h, '')) for h in headers]
                    else:
                        row = [str(record)]
                    table_data.append(row)

                table = Table(table_data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                ]))

                elements.append(table)

            # Footer
            footer_style = ParagraphStyle(
                'Footer',
                parent=styles['Normal'],
                fontSize=8,
                textColor=colors.grey,
                spaceAfter=10,
            )
            elements.append(Spacer(1, 0.3 * inch))
            elements.append(Paragraph(
                f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                footer_style
            ))

            doc.build(elements)
            output.seek(0)

            logger.info(f"PDF export created: {filename}")
            return output.getvalue()

        except ImportError:
            logger.error("reportlab not installed")
            return None
        except Exception as e:
            logger.error(f"PDF export error: {str(e)}")
            return None


class AttendanceExporter:
    """Attendance-specific export functionality"""

    @staticmethod
    def export_student_attendance(student, format='csv', start_date=None, end_date=None):
        """Export attendance records for a student"""
        from app1.models import Attendance

        queryset = Attendance.objects.filter(student=student)

        if start_date:
            queryset = queryset.filter(date__gte=start_date)
        if end_date:
            queryset = queryset.filter(date__lte=end_date)

        data = []
        for record in queryset:
            data.append({
                'Date': record.date.isoformat(),
                'Status': record.status,
                'Check-in Time': str(record.check_in_time) if record.check_in_time else 'N/A',
                'Check-out Time': str(record.check_out_time) if record.check_out_time else 'N/A',
            })

        filename = f"attendance_{student.roll_number}_{date.today().isoformat()}"

        if format == 'csv':
            return ExportService.export_to_csv(data, f"{filename}.csv")
        elif format == 'json':
            return ExportService.export_to_json(data, f"{filename}.json")
        elif format == 'xlsx':
            return ExportService.export_to_xlsx(data, f"{filename}.xlsx")
        elif format == 'pdf':
            return ExportService.export_to_pdf(
                data,
                f"{filename}.pdf",
                f"Attendance Report - {student.name}"
            )

        return None

    @staticmethod
    def export_class_attendance(class_name, format='csv', date_filter=None):
        """Export attendance for entire class"""
        from app1.models import Student, Attendance

        students = Student.objects.filter(student_class__icontains=class_name)
        data = []

        for student in students:
            queryset = Attendance.objects.filter(student=student)
            if date_filter:
                queryset = queryset.filter(date=date_filter)

            for record in queryset:
                data.append({
                    'Student': student.name,
                    'Roll Number': student.roll_number,
                    'Date': record.date.isoformat(),
                    'Status': record.status,
                    'Check-in': str(record.check_in_time) if record.check_in_time else 'N/A',
                })

        filename = f"class_attendance_{class_name}_{date.today().isoformat()}"

        if format == 'csv':
            return ExportService.export_to_csv(data, f"{filename}.csv")
        elif format == 'json':
            return ExportService.export_to_json(data, f"{filename}.json")
        elif format == 'xlsx':
            return ExportService.export_to_xlsx(data, f"{filename}.xlsx")
        elif format == 'pdf':
            return ExportService.export_to_pdf(
                data,
                f"{filename}.pdf",
                f"Class Attendance Report - {class_name}"
            )

        return None

    @staticmethod
    def export_analytics_report(analytics_data, format='json'):
        """Export analytics report"""
        filename = f"analytics_report_{date.today().isoformat()}"

        if format == 'json':
            return ExportService.export_to_json(analytics_data, f"{filename}.json")
        elif format == 'csv':
            # Convert analytics to flat structure for CSV
            flat_data = ExportService._flatten_dict(analytics_data)
            return ExportService.export_to_csv(flat_data, f"{filename}.csv")
        elif format == 'pdf':
            return ExportService.export_to_pdf(
                [analytics_data],
                f"{filename}.pdf",
                "Analytics Report"
            )

        return None

    @staticmethod
    def _flatten_dict(d, parent_key='', sep='_'):
        """Flatten nested dictionary"""
        items = []
        for k, v in d.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            if isinstance(v, dict):
                items.extend(ExportService._flatten_dict(v, new_key, sep=sep).items())
            elif isinstance(v, list):
                items.append((new_key, json.dumps(v)))
            else:
                items.append((new_key, v))
        return dict(items)


class TemplateExporter:
    """Export with custom templates"""

    CUSTOM_TEMPLATES = {}

    @classmethod
    def register_template(cls, template_id, template_func):
        """Register custom export template"""
        cls.CUSTOM_TEMPLATES[template_id] = template_func
        logger.info(f"Export template registered: {template_id}")

    @classmethod
    def export_with_template(cls, template_id, data):
        """Export using custom template"""
        if template_id in cls.CUSTOM_TEMPLATES:
            return cls.CUSTOM_TEMPLATES[template_id](data)
        return None

    @classmethod
    def list_templates(cls):
        """List available templates"""
        return list(cls.CUSTOM_TEMPLATES.keys())
